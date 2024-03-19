import openpyxl
from PIL import Image, ImageTk
from tkinter import messagebox
import formInputFile
import formLoginRegisterModule
import middleware

def register(username_val, password_val, window, self):
    # Define the header
    header = ['Username', 'Password', 'Status']
    
    # Create a new row with the user input
    new_row = [username_val, password_val, 'isLogin']

    # Append the new row to the Excel sheet
    workbook = openpyxl.load_workbook("CloudFileSystem.xlsx")
    sheet = workbook['user']
    sheet.append(new_row)
    workbook.save("CloudFileSystem.xlsx")
    middleware.middleware(self, window)
    # Melakukan sesuatu dengan nilai yang diperoleh
    messagebox.showinfo(title='Success', message=f'Registration Successful with username {username_val} and password {password_val}')
    
    
    
def login(username_val, password_val, window, self):
    # Load Workbook
    workbook = openpyxl.load_workbook("CloudFileSystem.xlsx")
    sheet = workbook['user']
    
    for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        if row[0] == username_val and row[1] == password_val:
            # Temukan sel status dan ubah nilainya
            status_cell = sheet.cell(row=row_idx, column=3)  # Kolom ke-3 adalah kolom status
            status_cell.value = 'isLogin'
            workbook.save("CloudFileSystem.xlsx")  # Simpan perubahan
            middleware.middleware(self, window)
            messagebox.showinfo(title='Success', message=f'Login is successful, welcome {username_val} to my application.')
            return True
        
    messagebox.showerror(title='Error', message='Login failed. Invalid username or password.')
    return False




def display_all_data():
    # Load workbook
    workbook = openpyxl.load_workbook("CloudFileSystem.xlsx")
    sheet = workbook['user']

    # List sheet
    for row in sheet.iter_rows(min_row=2, values_only=True):
        return row
    
def logout(self, window):
    confirm = messagebox.askyesno("Konfirmasi", "Anda yakin ingin logout?")
    
    if confirm:
        data = display_all_data()
        # Load Workbook
        workbook = openpyxl.load_workbook("CloudFileSystem.xlsx")
        sheet = workbook['user']
        
        # Mencari baris dengan username dan password yang sesuai
        for row in range(2, sheet.max_row + 1):
            username = sheet.cell(row=row, column=1).value
            password = sheet.cell(row=row, column=2).value
            status = sheet.cell(row=row, column=3).value
            
            if username == data[0] and password == data[1] and status == 'isLogin':
                # Ubah status menjadi 'isLogout'
                sheet.cell(row=row, column=3).value = 'isLogout'
                messagebox.showinfo(title='Success', message=f'Logout is successful, goodbye {data[0]}!')
                workbook.save("CloudFileSystem.xlsx")
                formLoginRegisterModule.show_form_login_register(self, window)
                return True
        
        # Jika username dan password tidak valid, atau pengguna sudah logout sebelumnya
        messagebox.showerror(title='Error', message='Logout failed. Invalid username or password or already logged out.')
        return False

    
        

